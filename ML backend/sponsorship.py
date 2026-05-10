"""H-1B sponsorship lookup and scoring for VisaRank.

Loads USCIS H-1B Employer Data Hub records into SQLite, fuzzy-matches a
LinkedIn-style company name against the dataset, and produces a 1-5 tier
score for "is this company a serious H-1B sponsor for new visa hires."

Public API (used by the Flask backend):
    get_company_features(name)               -> dict | None
    score_company(features)                  -> (int 1..5, str reason)
    score_with_cold_start(name, naics, state) -> dict (full verdict; uses
        sponsorship_model.py as a cold-start fallback when the lookup misses)

CLI:
    python sponsorship.py            # runs the demo bundle
    python sponsorship.py "Stripe"   # one-off lookup
"""

from __future__ import annotations

import csv
import os
import re
import sqlite3
import sys
from pathlib import Path
from typing import Optional

from rapidfuzz import fuzz, process

PROJECT_ROOT = Path(__file__).resolve().parent
DB_PATH = PROJECT_ROOT / "visarank.db"

# Tier 5 wants ~50 new visa hires per fiscal year of data we aggregated.
# Tier 4 wants ~10 per year. Tier 3 is anything positive.
TIER5_PER_YEAR = 50
TIER4_PER_YEAR = 10

FUZZY_CUTOFF = 80
FUZZY_TOP_K = 5

# How many recent fiscal years to aggregate for the user-facing answer.
RECENT_YEARS_WINDOW = 2


# ---------------------------------------------------------------------------
# Name normalization
# ---------------------------------------------------------------------------

_SUFFIX_RE = re.compile(
    r"\b(INC|INCORPORATED|LLC|L\s*L\s*C|CORP|CORPORATION|LTD|LIMITED|"
    r"LP|LLP|PLC|PC|PLLC|CO|COMPANY|HOLDINGS|GMBH|SA|NV|AG)\b\.?\s*$"
)
_PUNCT_RE = re.compile(r"[^\w\s]")
_WS_RE = re.compile(r"\s+")


def normalize_name(name: str) -> str:
    """Uppercase, strip punctuation, drop trailing legal suffixes."""
    if not name:
        return ""
    s = _PUNCT_RE.sub(" ", name.upper())
    s = _WS_RE.sub(" ", s).strip()
    # A name can carry several suffixes (e.g. "FOO INC LLC"); strip until stable.
    for _ in range(4):
        new = _SUFFIX_RE.sub("", s).strip()
        if new == s or not new:
            break
        s = new
    return s


# ---------------------------------------------------------------------------
# Source file discovery
# ---------------------------------------------------------------------------

def _find_source_file() -> Optional[Path]:
    candidates: list[Path] = []
    search_dirs = (
        PROJECT_ROOT,
        PROJECT_ROOT.parent,  # repo root (this module now lives in ML backend/)
        Path.home() / "Downloads",
    )
    for d in search_dirs:
        if not d.is_dir():
            continue
        for p in d.iterdir():
            if not p.is_file():
                continue
            n = p.name.lower()
            if p.suffix.lower() not in (".csv", ".xlsx", ".tsv"):
                continue
            if "h1b" in n or "h-1b" in n or "employer information" in n:
                candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


# ---------------------------------------------------------------------------
# Source file reading
# ---------------------------------------------------------------------------

# USCIS Employer Data Hub canonical column names.
NUMERIC_COLS = [
    "New Employment Approval",
    "New Employment Denial",
    "Continuation Approval",
    "Continuation Denial",
    "Change with Same Employer Approval",
    "Change with Same Employer Denial",
    "New Concurrent Approval",
    "New Concurrent Denial",
    "Change of Employer Approval",
    "Change of Employer Denial",
    "Amended Approval",
    "Amended Denial",
]

DB_NUMERIC_COLS = [
    "new_emp_approval",
    "new_emp_denial",
    "continuation_approval",
    "continuation_denial",
    "change_same_emp_approval",
    "change_same_emp_denial",
    "new_concurrent_approval",
    "new_concurrent_denial",
    "change_of_emp_approval",
    "change_of_emp_denial",
    "amended_approval",
    "amended_denial",
]


def _to_int(s) -> int:
    if s is None:
        return 0
    s = str(s).strip().replace(",", "")
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _detect_encoding(path: Path) -> str:
    with open(path, "rb") as f:
        head = f.read(4)
    if head.startswith(b"\xff\xfe") or head.startswith(b"\xfe\xff"):
        return "utf-16"
    return "utf-8-sig"


def _iter_csv_rows(path: Path):
    enc = _detect_encoding(path)
    with open(path, "r", encoding=enc, newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        # Strip whitespace from headers (the USCIS file has trailing spaces).
        reader.fieldnames = [
            (h or "").strip() for h in (reader.fieldnames or [])
        ]
        for row in reader:
            yield {(k or "").strip(): v for k, v in row.items()}


def _iter_xlsx_rows(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "xlsx source detected but openpyxl is not installed. "
            "Install it with: pip install openpyxl"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [(str(h).strip() if h is not None else "") for h in next(rows)]
    for r in rows:
        yield dict(zip(headers, r))


def _iter_source_rows(path: Path):
    if path.suffix.lower() == ".xlsx":
        yield from _iter_xlsx_rows(path)
    else:
        yield from _iter_csv_rows(path)


# ---------------------------------------------------------------------------
# DB: schema + ingest
# ---------------------------------------------------------------------------

# Bump this when the table layout changes; old DBs are dropped on first open.
SCHEMA_VERSION = "2"

SCHEMA = f"""
CREATE TABLE IF NOT EXISTS companies (
    normalized_name TEXT NOT NULL,
    fiscal_year INTEGER NOT NULL,
    display_name TEXT NOT NULL,
    naics_code TEXT NOT NULL DEFAULT '',
    state TEXT NOT NULL DEFAULT '',
    city TEXT NOT NULL DEFAULT '',
    zip TEXT NOT NULL DEFAULT '',
    {','.join(c + ' INTEGER NOT NULL DEFAULT 0' for c in DB_NUMERIC_COLS)},
    PRIMARY KEY (normalized_name, fiscal_year)
);
CREATE INDEX IF NOT EXISTS idx_companies_norm ON companies(normalized_name);
CREATE INDEX IF NOT EXISTS idx_companies_year ON companies(fiscal_year);
CREATE INDEX IF NOT EXISTS idx_companies_naics ON companies(naics_code);

CREATE TABLE IF NOT EXISTS meta (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);
"""


def _open_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    # If the on-disk schema predates the current version, wipe and let the
    # next ingest call rebuild it.
    cur = conn.execute("SELECT value FROM meta WHERE key = 'schema_version'")
    row = cur.fetchone()
    on_disk = row[0] if row else None
    if on_disk != SCHEMA_VERSION:
        with conn:
            conn.execute("DROP TABLE IF EXISTS companies")
            conn.executescript(SCHEMA)
            conn.execute(
                "INSERT OR REPLACE INTO meta(key, value) VALUES('schema_version', ?)",
                (SCHEMA_VERSION,),
            )
    return conn


def _db_has_data(conn: sqlite3.Connection) -> bool:
    cur = conn.execute("SELECT COUNT(*) FROM companies")
    return cur.fetchone()[0] > 0


def ingest(force: bool = False) -> int:
    """Populate the SQLite DB from the source file. Returns rows inserted.

    No-op (returns 0) if the DB already has data and `force` is False.
    """
    conn = _open_db()
    if _db_has_data(conn) and not force:
        return 0

    src = _find_source_file()
    if src is None:
        raise FileNotFoundError(
            "Could not find USCIS H-1B source file. Looked in project dir "
            f"({PROJECT_ROOT}) and ~/Downloads for files matching 'h1b' / "
            "'employer information'."
        )

    # Aggregate in memory by (normalized_name, fiscal_year). The raw file has
    # one row per (employer, FY, location) — multiple offices of the same
    # company show up as separate rows that we collapse here.
    from collections import Counter
    agg: dict[tuple[str, int], list] = {}
    display: dict[str, str] = {}
    # NAICS / state / city / zip vary across rows for the same normalized name
    # (different filing offices). Track frequency per (norm, fy) and pick the
    # mode at insert time.
    meta_counts: dict[tuple[str, int, str], Counter] = {}

    for row in _iter_source_rows(src):
        raw_name = (row.get("Employer (Petitioner) Name") or "").strip()
        if not raw_name:
            continue
        norm = normalize_name(raw_name)
        if not norm:
            continue
        fy = _to_int(row.get("Fiscal Year"))
        if fy <= 0:
            continue
        key = (norm, fy)
        if key not in agg:
            agg[key] = [0] * len(DB_NUMERIC_COLS)
        bucket = agg[key]
        for i, col in enumerate(NUMERIC_COLS):
            bucket[i] += _to_int(row.get(col))
        # NAICS in the raw file is "54 - Professional, Scientific, and Technical
        # Services"; keep just the leading code (or the whole string if no code).
        naics_raw = (row.get("Industry (NAICS) Code") or "").strip()
        naics = naics_raw.split(" - ")[0].strip() if naics_raw else ""
        state = (row.get("Petitioner State") or "").strip().upper()
        city = (row.get("Petitioner City") or "").strip().upper()
        zip_code = (row.get("Petitioner Zip Code") or "").strip()
        for field, value in (
            ("naics", naics), ("state", state), ("city", city), ("zip", zip_code)
        ):
            mc_key = (norm, fy, field)
            if mc_key not in meta_counts:
                meta_counts[mc_key] = Counter()
            if value:
                meta_counts[mc_key][value] += 1
        # Stash a representative display name; prefer the longest non-empty one
        # since the longest is usually the most official.
        prev = display.get(norm)
        if prev is None or len(raw_name) > len(prev):
            display[norm] = raw_name

    def _mode(norm: str, fy: int, field: str) -> str:
        c = meta_counts.get((norm, fy, field))
        if not c:
            return ""
        return c.most_common(1)[0][0]

    cols = [
        "normalized_name", "fiscal_year", "display_name",
        "naics_code", "state", "city", "zip",
        *DB_NUMERIC_COLS,
    ]
    placeholders = ",".join(["?"] * len(cols))
    insert_sql = (
        f"INSERT OR REPLACE INTO companies ({','.join(cols)}) "
        f"VALUES ({placeholders})"
    )
    rows_to_insert = [
        (
            norm, fy, display[norm],
            _mode(norm, fy, "naics"),
            _mode(norm, fy, "state"),
            _mode(norm, fy, "city"),
            _mode(norm, fy, "zip"),
            *counts,
        )
        for (norm, fy), counts in agg.items()
    ]
    with conn:
        conn.execute("DELETE FROM companies")
        conn.executemany(insert_sql, rows_to_insert)
        conn.execute(
            "INSERT OR REPLACE INTO meta(key, value) VALUES('source', ?)",
            (str(src),),
        )
        conn.execute(
            "INSERT OR REPLACE INTO meta(key, value) VALUES('schema_version', ?)",
            (SCHEMA_VERSION,),
        )
    conn.close()
    return len(rows_to_insert)


# ---------------------------------------------------------------------------
# Lookup + scoring
# ---------------------------------------------------------------------------

_NAME_INDEX_CACHE: Optional[list[str]] = None


def _all_normalized_names(conn: sqlite3.Connection) -> list[str]:
    global _NAME_INDEX_CACHE
    if _NAME_INDEX_CACHE is None:
        cur = conn.execute("SELECT DISTINCT normalized_name FROM companies")
        _NAME_INDEX_CACHE = [r[0] for r in cur.fetchall()]
    return _NAME_INDEX_CACHE


def _fetch_rows(conn: sqlite3.Connection, normalized_name: str):
    cur = conn.execute(
        f"SELECT fiscal_year, display_name, {','.join(DB_NUMERIC_COLS)} "
        f"FROM companies WHERE normalized_name = ? ORDER BY fiscal_year DESC",
        (normalized_name,),
    )
    return cur.fetchall()


def _ensure_loaded() -> sqlite3.Connection:
    conn = _open_db()
    if not _db_has_data(conn):
        conn.close()
        ingest(force=False)
        conn = _open_db()
    return conn


def _resolve_match(conn: sqlite3.Connection, raw_name: str):
    """Resolve a free-form name to (normalized_names, match_score, label).

    Three stages, in order:
        1. Exact normalized match.
        2. Prefix aggregation: every "<query> *" subsidiary is rolled up.
           Handles 'Amazon' -> all AMAZON COM/AWS/RETAIL/... entities.
        3. Token-subset fuzzy: ``token_set_ratio`` plus a hard requirement
           that every >=3-char query token appears as an exact token in the
           candidate. Blocks 'Meta'->METALS / 'ABC Random'->ABC ACCOUNTING.
    """
    norm = normalize_name(raw_name)
    if not norm:
        return None

    rows = _fetch_rows(conn, norm)
    if rows:
        return [norm], 100.0, rows[0][1]

    cur = conn.execute(
        "SELECT DISTINCT normalized_name FROM companies "
        "WHERE normalized_name LIKE ?",
        (norm + " %",),
    )
    prefix_hits = [r[0] for r in cur.fetchall()]
    if prefix_hits:
        label = (
            raw_name.strip()
            if len(prefix_hits) > 1
            else _fetch_rows(conn, prefix_hits[0])[0][1]
        )
        if len(prefix_hits) > 1:
            label = f"{label} ({len(prefix_hits)} subsidiaries)"
        return prefix_hits, 95.0, label

    names = _all_normalized_names(conn)
    query_tokens = {t for t in norm.split() if len(t) >= 3}
    if not query_tokens:
        return None
    results = process.extract(
        norm, names, scorer=fuzz.token_set_ratio,
        limit=20, score_cutoff=FUZZY_CUTOFF,
    )
    for matched, score, _ in results:
        cand_tokens = set(matched.split())
        if query_tokens.issubset(cand_tokens):
            label = _fetch_rows(conn, matched)[0][1]
            return [matched], float(score), label
    return None


def get_company_features(raw_name: str) -> Optional[dict]:
    """Match a free-form company name to the USCIS dataset and return
    aggregated features over the most recent ``RECENT_YEARS_WINDOW`` fiscal
    years.

    Returns None if no row clears the fuzzy match threshold.
    """
    if not raw_name or not raw_name.strip():
        return None

    conn = _ensure_loaded()
    try:
        resolved = _resolve_match(conn, raw_name)
        if resolved is None:
            return None
        matched_names, match_score, label = resolved

        rows = []
        for n in matched_names:
            rows.extend(_fetch_rows(conn, n))
        if not rows:
            return None

        # Aggregate over the most recent N fiscal years available, summing
        # across all matched normalized names (subsidiaries).
        all_years = sorted({r[0] for r in rows}, reverse=True)
        window_years = all_years[:RECENT_YEARS_WINDOW]
        window_rows = [r for r in rows if r[0] in window_years]

        per_year: dict[int, dict] = {}
        for r in window_rows:
            fy = r[0]
            counts = dict(zip(DB_NUMERIC_COLS, r[2:]))
            if fy not in per_year:
                per_year[fy] = {c: 0 for c in DB_NUMERIC_COLS}
            for c, v in counts.items():
                per_year[fy][c] += v

        def total(col: str) -> int:
            return sum(per_year[y][col] for y in per_year)

        new_visa_hires = total("new_emp_approval") + total("change_of_emp_approval")
        continuations = total("continuation_approval")
        amendments = total("amended_approval")
        change_same = total("change_same_emp_approval")
        new_concurrent = total("new_concurrent_approval")
        all_approvals = sum(
            total(c) for c in DB_NUMERIC_COLS if c.endswith("_approval")
        )

        # Trend: latest year's new_visa_hires vs prior year's, if both exist.
        trend = None
        if len(window_years) >= 2:
            latest, prior = window_years[0], window_years[1]
            latest_new = (
                per_year[latest]["new_emp_approval"]
                + per_year[latest]["change_of_emp_approval"]
            )
            prior_new = (
                per_year[prior]["new_emp_approval"]
                + per_year[prior]["change_of_emp_approval"]
            )
            if prior_new == 0 and latest_new == 0:
                trend = "flat"
            elif prior_new == 0:
                trend = "up"
            elif latest_new >= prior_new * 1.1:
                trend = "up"
            elif latest_new <= prior_new * 0.9:
                trend = "down"
            else:
                trend = "flat"

        return {
            "matched_name": label,
            "normalized_names": matched_names,
            "match_score": float(match_score),
            "year_window": len(window_years),
            "years": window_years,
            "new_visa_hires": new_visa_hires,
            "continuations": continuations,
            "amendments": amendments,
            "change_with_same_employer": change_same,
            "new_concurrent": new_concurrent,
            "total_approvals": all_approvals,
            "trend": trend,
            "per_year": per_year,
        }
    finally:
        conn.close()


def score_company(features: Optional[dict]) -> tuple[int, str]:
    """Map an aggregated-features dict to a (1..5, reason) tuple.

    Tier 5: heavy new-hire sponsor, trending stable or up.
    Tier 4: solid new-hire sponsor.
    Tier 3: some new-hire sponsorship, modest volume.
    Tier 2: H-1B activity exists but no new visa hires (renewal-only).
    Tier 1: no H-1B activity on record / no company match.
    """
    if features is None:
        return 1, "No matching company found in the USCIS H-1B records."

    new_hires = features["new_visa_hires"]
    total_appr = features["total_approvals"]
    yrs = features["year_window"] or 1
    trend = features.get("trend")
    name = features["matched_name"]
    window_phrase = f"in FY{features['years'][-1]}" if yrs == 1 else f"in the last {yrs} fiscal years"

    if total_appr == 0:
        return 1, f"{name} has no H-1B approvals on record."

    if new_hires == 0:
        non_new = (
            features["continuations"]
            + features["amendments"]
            + features["change_with_same_employer"]
        )
        return 2, (
            f"{name} filed {non_new} H-1B renewals/amendments {window_phrase} "
            f"but zero new visa hires."
        )

    trend_clause = ""
    if trend == "up":
        trend_clause = ", trending up"
    elif trend == "down":
        trend_clause = ", trending down"
    elif trend == "flat":
        trend_clause = ", trending flat"

    t5 = TIER5_PER_YEAR * yrs
    t4 = TIER4_PER_YEAR * yrs

    if new_hires >= t5 and trend != "down":
        return 5, (
            f"{name} sponsored {new_hires} new H-1B workers "
            f"{window_phrase}{trend_clause}."
        )
    if new_hires >= t5 and trend == "down":
        # Heavy sponsor but declining — demote to tier 4.
        return 4, (
            f"{name} sponsored {new_hires} new H-1B workers "
            f"{window_phrase}{trend_clause}."
        )
    if new_hires >= t4:
        return 4, (
            f"{name} sponsored {new_hires} new H-1B workers "
            f"{window_phrase}{trend_clause}."
        )
    return 3, (
        f"{name} sponsored {new_hires} new H-1B workers "
        f"{window_phrase}{trend_clause}."
    )


# ---------------------------------------------------------------------------
# Cold-start fallback (sponsorship_model.py)
# ---------------------------------------------------------------------------

def _try_predict(
    name: str,
    naics: Optional[str],
    state: Optional[str],
    description: Optional[str] = None,
    title: Optional[str] = None,
) -> Optional[dict]:
    """Return the full predict_with_details dict, or None if the model
    isn't available. Lazy-imported so the lookup-only path stays free of
    sklearn/pandas.
    """
    try:
        from sponsorship_model import predict_with_details
        return predict_with_details(
            name, naics=naics, state=state,
            description=description, title=title,
        )
    except (ImportError, FileNotFoundError):
        return None
    except Exception:
        return None


def _try_posting(title: Optional[str], description: Optional[str]) -> Optional[dict]:
    """Run posting_filter.analyze_posting if any text is provided."""
    if not (title or description):
        return None
    try:
        from posting_filter import analyze_posting
        return analyze_posting(title or "", description or "")
    except ImportError:
        return None
    except Exception:
        return None


def _apply_posting_override(
    base_tier: int,
    base_reason: str,
    posting: dict,
    *,
    source: str,
) -> tuple[int, str, bool]:
    """Combine the company-level tier with the posting-level signal.

    Asymmetric on purpose: ``hard_no`` can collapse any tier to 1 (a posting
    that excludes non-citizens is dispositive regardless of company). But a
    ``hard_yes`` from a cold-start company can't push past tier 3, since we
    have no track record to back up the posting's claim.

    Returns (tier, reason, override_applied).
    """
    cat = posting["category"]
    evidence = posting.get("evidence") or []
    quote = evidence[0] if evidence else ""
    quote_clause = f' ("{quote}")' if quote else ""

    if cat == "hard_no":
        if source == "uscis":
            # Company has real H1B history, but THIS posting explicitly closes
            # the door. Surface both facts so the user understands the conflict.
            reason = (
                f"Despite past H-1B history — {base_reason.rstrip('.')} — "
                f"this posting explicitly excludes sponsorship{quote_clause}. "
                f"Do not apply expecting sponsorship for this role."
            )
        else:
            reason = f"Posting explicitly excludes sponsorship{quote_clause}."
        return 1, reason, True

    if cat == "soft_no":
        new_tier = max(1, base_tier - 1)
        reason = (
            f"{base_reason} This posting hedges on sponsorship{quote_clause}."
        )
        return new_tier, reason, True

    if cat == "hard_yes":
        # Cold-start cap at 3 (no track record); USCIS-matched can climb to 5.
        cap = 5 if source == "uscis" else 3
        new_tier = min(cap, base_tier + 1)
        if new_tier == base_tier:
            return base_tier, base_reason, False
        reason = (
            f"{base_reason} Posting explicitly offers sponsorship: {quote}"
            if quote else f"{base_reason} Posting explicitly offers sponsorship."
        )
        return new_tier, reason, True

    if cat == "soft_yes":
        cap = 5 if source == "uscis" else 3
        new_tier = min(cap, base_tier + 1)
        if new_tier == base_tier:
            return base_tier, base_reason, False
        reason = (
            f"{base_reason} Posting open to sponsorship: {quote}"
            if quote else f"{base_reason} Posting open to sponsorship."
        )
        return new_tier, reason, True

    # neutral
    return base_tier, base_reason, False


def score_with_cold_start(
    name: str,
    naics: Optional[str] = None,
    state: Optional[str] = None,
    description: Optional[str] = None,
    title: Optional[str] = None,
) -> dict:
    """Full pipeline: USCIS lookup first, fall back to LR predictor.

    The `description` / `title` are used by the LR model to impute a NAICS
    code via the keyword dictionary (sponsorship_model.NAICS_KEYWORDS) when
    the caller doesn't pass an explicit `naics`. The lookup path ignores
    them.

    Returns a dict the Flask layer can hand to the extension directly:
        tier        -- 1..5
        reason      -- one-line human string
        source      -- 'uscis' | 'model' | 'none'
        features    -- the lookup dict if matched, else None
        prediction  -- predicted probability if the model was used, else None
        match_score -- fuzzy match confidence if matched, else None
    """
    posting = _try_posting(title, description)

    feats = get_company_features(name)
    if feats is not None:
        tier, reason = score_company(feats)
        details = _try_predict(name, naics, state, description, title)
        override_applied = False
        if posting is not None:
            tier, reason, override_applied = _apply_posting_override(
                tier, reason, posting, source="uscis",
            )
        return {
            "tier": tier,
            "reason": reason,
            "source": "uscis",
            "features": feats,
            "prediction": details["p_sponsors"] if details else None,
            "model_details": details,
            "match_score": feats["match_score"],
            "posting_signal": posting,
            "override_applied": override_applied,
        }

    details = _try_predict(name, naics, state, description, title)
    if details is None:
        # No model and no lookup -- but a kill-switch posting is still
        # dispositive on its own ("not in our DB but says US-citizens-only"
        # is still a clear "don't bother").
        tier, reason = 1, "No matching company found in the USCIS H-1B records."
        override_applied = False
        if posting is not None:
            tier, reason, override_applied = _apply_posting_override(
                tier, reason, posting, source="none",
            )
        return {
            "tier": tier,
            "reason": reason,
            "source": "none",
            "features": None,
            "prediction": None,
            "model_details": None,
            "match_score": None,
            "posting_signal": posting,
            "override_applied": override_applied,
        }

    p = details["p_sponsors"]
    # The LR's held-out AUC is ~0.57 — modest. Cap cold-start verdicts at
    # tier 3 so we don't overclaim from a weak signal.
    if p >= 0.60:
        tier = 3
        reason = (
            f"No direct USCIS record; predictor estimates likely sponsor "
            f"(P={p:.2f}) from industry and name profile."
        )
    elif p >= 0.45:
        tier = 2
        reason = (
            f"No direct USCIS record; predictor is uncertain (P={p:.2f})."
        )
    else:
        tier = 1
        reason = (
            f"No direct USCIS record; predictor estimates unlikely sponsor "
            f"(P={p:.2f})."
        )

    override_applied = False
    if posting is not None:
        tier, reason, override_applied = _apply_posting_override(
            tier, reason, posting, source="model",
        )

    return {
        "tier": tier,
        "reason": reason,
        "source": "model",
        "features": None,
        "prediction": p,
        "model_details": details,
        "match_score": None,
        "posting_signal": posting,
        "override_applied": override_applied,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

DEMO_NAMES = [
    "Amazon",
    "Google",
    "Microsoft",
    "Apple",
    "Meta",
    "Stripe",
    "ABC Random LLC",
]


def _format_verdict(query: str, v: dict) -> str:
    lines = [f"\nQuery: {query!r}  [source={v['source']}]"]
    feats = v.get("features")
    md = v.get("model_details")
    if feats is not None:
        lines.append(f"  matched           : {feats['matched_name']}")
        lines.append(f"  name-match conf   : {feats['match_score']:.0f}%")
        lines.append(f"  years             : {feats['years']}")
        lines.append(f"  new H-1B hires    : {feats['new_visa_hires']}")
        lines.append(f"  continuations     : {feats['continuations']}")
        lines.append(f"  amendments        : {feats['amendments']}")
        lines.append(f"  total approvals   : {feats['total_approvals']}")
        if feats.get("trend"):
            lines.append(f"  trend             : {feats['trend']}")
        if v.get("prediction") is not None:
            lines.append(
                f"  model P(sponsors) : {v['prediction']:.3f}  "
                f"(independent LR estimate; data above is authoritative)"
            )
    elif v["prediction"] is not None:
        lines.append(
            f"  model P(sponsors) : {v['prediction']:.3f}  "
            f"(cold-start LR -- no USCIS record for this company)"
        )
        if md:
            tag = (
                f"NAICS {md['naics_used']} ({md['naics_source']})"
            )
            if md.get("naics_evidence"):
                tag += f" via [{', '.join(md['naics_evidence'])}]"
            lines.append(f"  industry inferred : {tag}")
            lines.append(f"  state used        : {md['state_used']}")
    else:
        lines.append("  (no match, no model)")

    posting = v.get("posting_signal")
    if posting is not None and posting.get("category") != "neutral":
        ev = (posting.get("evidence") or [""])[0]
        flag = "OVERRIDE" if v.get("override_applied") else "noted"
        lines.append(
            f"  posting signal    : {posting['category']:<9} [{flag}]"
        )
        if ev:
            lines.append(f"    evidence        : {ev}")

    lines.append(f"  TIER {v['tier']} -- {v['reason']}")
    return "\n".join(lines)


def main(argv: list[str]) -> int:
    if len(argv) >= 2 and argv[1] in ("--load", "--reload"):
        n = ingest(force=True)
        print(f"Loaded {n} (company, fiscal_year) rows into {DB_PATH}")
        return 0

    # Strip optional flags (--state X, --jd X, --title X) out of argv; the
    # remainder is the positional company-name list. We're hand-parsing here
    # rather than reaching for argparse so the bare `python sponsorship.py
    # "Stripe" "Amazon" ...` form keeps working unchanged.
    args = argv[1:]
    state = jd = title = None
    positional: list[str] = []
    i = 0
    while i < len(args):
        a = args[i]
        if a == "--state" and i + 1 < len(args):
            state = args[i + 1]
            i += 2
        elif a in ("--jd", "--description") and i + 1 < len(args):
            jd = args[i + 1]
            i += 2
        elif a == "--title" and i + 1 < len(args):
            title = args[i + 1]
            i += 2
        else:
            positional.append(a)
            i += 1

    n = ingest(force=False)
    if n:
        print(f"[ingest] populated DB with {n} (company, fiscal_year) rows")

    queries = positional or DEMO_NAMES
    for q in queries:
        v = score_with_cold_start(q, state=state, description=jd, title=title)
        print(_format_verdict(q, v))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
