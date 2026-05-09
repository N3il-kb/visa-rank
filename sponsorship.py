"""H-1B sponsorship lookup and scoring for VisaRank.

Loads USCIS H-1B Employer Data Hub records into SQLite, fuzzy-matches a
LinkedIn-style company name against the dataset, and produces a 1-5 tier
score for "is this company a serious H-1B sponsor for new visa hires."

Public API (used by the Flask backend):
    get_company_features(name) -> dict | None
    score_company(features)    -> (int 1..5, str reason)

CLI:
    python sponsorship.py            # runs the demo bundle
    python sponsorship.py "Stripe"   # one-off lookup
"""

from __future__ import annotations

import csv
import os
import re
import sqlite3
import sys
from pathlib import Path
from typing import Optional

from rapidfuzz import fuzz, process

PROJECT_ROOT = Path(__file__).resolve().parent
DB_PATH = PROJECT_ROOT / "visarank.db"

# Tier 5 wants ~50 new visa hires per fiscal year of data we aggregated.
# Tier 4 wants ~10 per year. Tier 3 is anything positive.
TIER5_PER_YEAR = 50
TIER4_PER_YEAR = 10

FUZZY_CUTOFF = 80
FUZZY_TOP_K = 5

# How many recent fiscal years to aggregate for the user-facing answer.
RECENT_YEARS_WINDOW = 2


# ---------------------------------------------------------------------------
# Name normalization
# ---------------------------------------------------------------------------

_SUFFIX_RE = re.compile(
    r"\b(INC|INCORPORATED|LLC|L\s*L\s*C|CORP|CORPORATION|LTD|LIMITED|"
    r"LP|LLP|PLC|PC|PLLC|CO|COMPANY|HOLDINGS|GMBH|SA|NV|AG)\b\.?\s*$"
)
_PUNCT_RE = re.compile(r"[^\w\s]")
_WS_RE = re.compile(r"\s+")


def normalize_name(name: str) -> str:
    """Uppercase, strip punctuation, drop trailing legal suffixes."""
    if not name:
        return ""
    s = _PUNCT_RE.sub(" ", name.upper())
    s = _WS_RE.sub(" ", s).strip()
    # A name can carry several suffixes (e.g. "FOO INC LLC"); strip until stable.
    for _ in range(4):
        new = _SUFFIX_RE.sub("", s).strip()
        if new == s or not new:
            break
        s = new
    return s


# ---------------------------------------------------------------------------
# Source file discovery
# ---------------------------------------------------------------------------

def _find_source_file() -> Optional[Path]:
    candidates: list[Path] = []
    for d in (PROJECT_ROOT, Path.home() / "Downloads"):
        if not d.is_dir():
            continue
        for p in d.iterdir():
            if not p.is_file():
                continue
            n = p.name.lower()
            if p.suffix.lower() not in (".csv", ".xlsx", ".tsv"):
                continue
            if "h1b" in n or "h-1b" in n or "employer information" in n:
                candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


# ---------------------------------------------------------------------------
# Source file reading
# ---------------------------------------------------------------------------

# USCIS Employer Data Hub canonical column names.
NUMERIC_COLS = [
    "New Employment Approval",
    "New Employment Denial",
    "Continuation Approval",
    "Continuation Denial",
    "Change with Same Employer Approval",
    "Change with Same Employer Denial",
    "New Concurrent Approval",
    "New Concurrent Denial",
    "Change of Employer Approval",
    "Change of Employer Denial",
    "Amended Approval",
    "Amended Denial",
]

DB_NUMERIC_COLS = [
    "new_emp_approval",
    "new_emp_denial",
    "continuation_approval",
    "continuation_denial",
    "change_same_emp_approval",
    "change_same_emp_denial",
    "new_concurrent_approval",
    "new_concurrent_denial",
    "change_of_emp_approval",
    "change_of_emp_denial",
    "amended_approval",
    "amended_denial",
]


def _to_int(s) -> int:
    if s is None:
        return 0
    s = str(s).strip().replace(",", "")
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _detect_encoding(path: Path) -> str:
    with open(path, "rb") as f:
        head = f.read(4)
    if head.startswith(b"\xff\xfe") or head.startswith(b"\xfe\xff"):
        return "utf-16"
    return "utf-8-sig"


def _iter_csv_rows(path: Path):
    enc = _detect_encoding(path)
    with open(path, "r", encoding=enc, newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        # Strip whitespace from headers (the USCIS file has trailing spaces).
        reader.fieldnames = [
            (h or "").strip() for h in (reader.fieldnames or [])
        ]
        for row in reader:
            yield {(k or "").strip(): v for k, v in row.items()}


def _iter_xlsx_rows(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise RuntimeError(
            "xlsx source detected but openpyxl is not installed. "
            "Install it with: pip install openpyxl"
        ) from e
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [(str(h).strip() if h is not None else "") for h in next(rows)]
    for r in rows:
        yield dict(zip(headers, r))


def _iter_source_rows(path: Path):
    if path.suffix.lower() == ".xlsx":
        yield from _iter_xlsx_rows(path)
    else:
        yield from _iter_csv_rows(path)


# ---------------------------------------------------------------------------
# DB: schema + ingest
# ---------------------------------------------------------------------------

SCHEMA = f"""
CREATE TABLE IF NOT EXISTS companies (
    normalized_name TEXT NOT NULL,
    fiscal_year INTEGER NOT NULL,
    display_name TEXT NOT NULL,
    {','.join(c + ' INTEGER NOT NULL DEFAULT 0' for c in DB_NUMERIC_COLS)},
    PRIMARY KEY (normalized_name, fiscal_year)
);
CREATE INDEX IF NOT EXISTS idx_companies_norm ON companies(normalized_name);
CREATE INDEX IF NOT EXISTS idx_companies_year ON companies(fiscal_year);

CREATE TABLE IF NOT EXISTS meta (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);
"""


def _open_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    return conn


def _db_has_data(conn: sqlite3.Connection) -> bool:
    cur = conn.execute("SELECT COUNT(*) FROM companies")
    return cur.fetchone()[0] > 0


def ingest(force: bool = False) -> int:
    """Populate the SQLite DB from the source file. Returns rows inserted.

    No-op (returns 0) if the DB already has data and `force` is False.
    """
    conn = _open_db()
    if _db_has_data(conn) and not force:
        return 0

    src = _find_source_file()
    if src is None:
        raise FileNotFoundError(
            "Could not find USCIS H-1B source file. Looked in project dir "
            f"({PROJECT_ROOT}) and ~/Downloads for files matching 'h1b' / "
            "'employer information'."
        )

    # Aggregate in memory by (normalized_name, fiscal_year). The raw file has
    # one row per (employer, FY, location) — multiple offices of the same
    # company show up as separate rows that we collapse here.
    agg: dict[tuple[str, int], list] = {}
    display: dict[str, str] = {}

    for row in _iter_source_rows(src):
        raw_name = (row.get("Employer (Petitioner) Name") or "").strip()
        if not raw_name:
            continue
        norm = normalize_name(raw_name)
        if not norm:
            continue
        fy = _to_int(row.get("Fiscal Year"))
        if fy <= 0:
            continue
        key = (norm, fy)
        if key not in agg:
            agg[key] = [0] * len(DB_NUMERIC_COLS)
        bucket = agg[key]
        for i, col in enumerate(NUMERIC_COLS):
            bucket[i] += _to_int(row.get(col))
        # Stash a representative display name; prefer the longest non-empty one
        # since the longest is usually the most official.
        prev = display.get(norm)
        if prev is None or len(raw_name) > len(prev):
            display[norm] = raw_name

    placeholders = ",".join(["?"] * (3 + len(DB_NUMERIC_COLS)))
    insert_sql = (
        f"INSERT OR REPLACE INTO companies "
        f"(normalized_name, fiscal_year, display_name, "
        f"{','.join(DB_NUMERIC_COLS)}) VALUES ({placeholders})"
    )
    rows_to_insert = [
        (norm, fy, display[norm], *counts)
        for (norm, fy), counts in agg.items()
    ]
    with conn:
        conn.execute("DELETE FROM companies")
        conn.executemany(insert_sql, rows_to_insert)
        conn.execute(
            "INSERT OR REPLACE INTO meta(key, value) VALUES('source', ?)",
            (str(src),),
        )
    conn.close()
    return len(rows_to_insert)


# ---------------------------------------------------------------------------
# Lookup + scoring
# ---------------------------------------------------------------------------

_NAME_INDEX_CACHE: Optional[list[str]] = None


def _all_normalized_names(conn: sqlite3.Connection) -> list[str]:
    global _NAME_INDEX_CACHE
    if _NAME_INDEX_CACHE is None:
        cur = conn.execute("SELECT DISTINCT normalized_name FROM companies")
        _NAME_INDEX_CACHE = [r[0] for r in cur.fetchall()]
    return _NAME_INDEX_CACHE


def _fetch_rows(conn: sqlite3.Connection, normalized_name: str):
    cur = conn.execute(
        f"SELECT fiscal_year, display_name, {','.join(DB_NUMERIC_COLS)} "
        f"FROM companies WHERE normalized_name = ? ORDER BY fiscal_year DESC",
        (normalized_name,),
    )
    return cur.fetchall()


def _ensure_loaded() -> sqlite3.Connection:
    conn = _open_db()
    if not _db_has_data(conn):
        conn.close()
        ingest(force=False)
        conn = _open_db()
    return conn


def _resolve_match(conn: sqlite3.Connection, raw_name: str):
    """Resolve a free-form name to (normalized_names, match_score, label).

    Three stages, in order:
        1. Exact normalized match.
        2. Prefix aggregation: every "<query> *" subsidiary is rolled up.
           Handles 'Amazon' -> all AMAZON COM/AWS/RETAIL/... entities.
        3. Token-subset fuzzy: ``token_set_ratio`` plus a hard requirement
           that every >=3-char query token appears as an exact token in the
           candidate. Blocks 'Meta'->METALS / 'ABC Random'->ABC ACCOUNTING.
    """
    norm = normalize_name(raw_name)
    if not norm:
        return None

    rows = _fetch_rows(conn, norm)
    if rows:
        return [norm], 100.0, rows[0][1]

    cur = conn.execute(
        "SELECT DISTINCT normalized_name FROM companies "
        "WHERE normalized_name LIKE ?",
        (norm + " %",),
    )
    prefix_hits = [r[0] for r in cur.fetchall()]
    if prefix_hits:
        label = (
            raw_name.strip()
            if len(prefix_hits) > 1
            else _fetch_rows(conn, prefix_hits[0])[0][1]
        )
        if len(prefix_hits) > 1:
            label = f"{label} ({len(prefix_hits)} subsidiaries)"
        return prefix_hits, 95.0, label

    names = _all_normalized_names(conn)
    query_tokens = {t for t in norm.split() if len(t) >= 3}
    if not query_tokens:
        return None
    results = process.extract(
        norm, names, scorer=fuzz.token_set_ratio,
        limit=20, score_cutoff=FUZZY_CUTOFF,
    )
    for matched, score, _ in results:
        cand_tokens = set(matched.split())
        if query_tokens.issubset(cand_tokens):
            label = _fetch_rows(conn, matched)[0][1]
            return [matched], float(score), label
    return None


def get_company_features(raw_name: str) -> Optional[dict]:
    """Match a free-form company name to the USCIS dataset and return
    aggregated features over the most recent ``RECENT_YEARS_WINDOW`` fiscal
    years.

    Returns None if no row clears the fuzzy match threshold.
    """
    if not raw_name or not raw_name.strip():
        return None

    conn = _ensure_loaded()
    try:
        resolved = _resolve_match(conn, raw_name)
        if resolved is None:
            return None
        matched_names, match_score, label = resolved

        rows = []
        for n in matched_names:
            rows.extend(_fetch_rows(conn, n))
        if not rows:
            return None

        # Aggregate over the most recent N fiscal years available, summing
        # across all matched normalized names (subsidiaries).
        all_years = sorted({r[0] for r in rows}, reverse=True)
        window_years = all_years[:RECENT_YEARS_WINDOW]
        window_rows = [r for r in rows if r[0] in window_years]

        per_year: dict[int, dict] = {}
        for r in window_rows:
            fy = r[0]
            counts = dict(zip(DB_NUMERIC_COLS, r[2:]))
            if fy not in per_year:
                per_year[fy] = {c: 0 for c in DB_NUMERIC_COLS}
            for c, v in counts.items():
                per_year[fy][c] += v

        def total(col: str) -> int:
            return sum(per_year[y][col] for y in per_year)

        new_visa_hires = total("new_emp_approval") + total("change_of_emp_approval")
        continuations = total("continuation_approval")
        amendments = total("amended_approval")
        change_same = total("change_same_emp_approval")
        new_concurrent = total("new_concurrent_approval")
        all_approvals = sum(
            total(c) for c in DB_NUMERIC_COLS if c.endswith("_approval")
        )

        # Trend: latest year's new_visa_hires vs prior year's, if both exist.
        trend = None
        if len(window_years) >= 2:
            latest, prior = window_years[0], window_years[1]
            latest_new = (
                per_year[latest]["new_emp_approval"]
                + per_year[latest]["change_of_emp_approval"]
            )
            prior_new = (
                per_year[prior]["new_emp_approval"]
                + per_year[prior]["change_of_emp_approval"]
            )
            if prior_new == 0 and latest_new == 0:
                trend = "flat"
            elif prior_new == 0:
                trend = "up"
            elif latest_new >= prior_new * 1.1:
                trend = "up"
            elif latest_new <= prior_new * 0.9:
                trend = "down"
            else:
                trend = "flat"

        return {
            "matched_name": label,
            "normalized_names": matched_names,
            "match_score": float(match_score),
            "year_window": len(window_years),
            "years": window_years,
            "new_visa_hires": new_visa_hires,
            "continuations": continuations,
            "amendments": amendments,
            "change_with_same_employer": change_same,
            "new_concurrent": new_concurrent,
            "total_approvals": all_approvals,
            "trend": trend,
            "per_year": per_year,
        }
    finally:
        conn.close()


def score_company(features: Optional[dict]) -> tuple[int, str]:
    """Map an aggregated-features dict to a (1..5, reason) tuple.

    Tier 5: heavy new-hire sponsor, trending stable or up.
    Tier 4: solid new-hire sponsor.
    Tier 3: some new-hire sponsorship, modest volume.
    Tier 2: H-1B activity exists but no new visa hires (renewal-only).
    Tier 1: no H-1B activity on record / no company match.
    """
    if features is None:
        return 1, "No matching company found in the USCIS H-1B records."

    new_hires = features["new_visa_hires"]
    total_appr = features["total_approvals"]
    yrs = features["year_window"] or 1
    trend = features.get("trend")
    name = features["matched_name"]
    window_phrase = f"in FY{features['years'][-1]}" if yrs == 1 else f"in the last {yrs} fiscal years"

    if total_appr == 0:
        return 1, f"{name} has no H-1B approvals on record."

    if new_hires == 0:
        non_new = (
            features["continuations"]
            + features["amendments"]
            + features["change_with_same_employer"]
        )
        return 2, (
            f"{name} filed {non_new} H-1B renewals/amendments {window_phrase} "
            f"but zero new visa hires."
        )

    trend_clause = ""
    if trend == "up":
        trend_clause = ", trending up"
    elif trend == "down":
        trend_clause = ", trending down"
    elif trend == "flat":
        trend_clause = ", trending flat"

    t5 = TIER5_PER_YEAR * yrs
    t4 = TIER4_PER_YEAR * yrs

    if new_hires >= t5 and trend != "down":
        return 5, (
            f"{name} sponsored {new_hires} new H-1B workers "
            f"{window_phrase}{trend_clause}."
        )
    if new_hires >= t5 and trend == "down":
        # Heavy sponsor but declining — demote to tier 4.
        return 4, (
            f"{name} sponsored {new_hires} new H-1B workers "
            f"{window_phrase}{trend_clause}."
        )
    if new_hires >= t4:
        return 4, (
            f"{name} sponsored {new_hires} new H-1B workers "
            f"{window_phrase}{trend_clause}."
        )
    return 3, (
        f"{name} sponsored {new_hires} new H-1B workers "
        f"{window_phrase}{trend_clause}."
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

DEMO_NAMES = [
    "Amazon",
    "Google",
    "Microsoft",
    "Apple",
    "Meta",
    "Stripe",
    "ABC Random LLC",
]


def _format_result(query: str, features: Optional[dict], tier: int, reason: str) -> str:
    lines = [f"\nQuery: {query!r}"]
    if features is None:
        lines.append("  (no match)")
    else:
        lines.append(f"  matched     : {features['matched_name']}")
        lines.append(f"  match score : {features['match_score']:.1f}")
        lines.append(f"  years       : {features['years']}")
        lines.append(f"  new H-1B    : {features['new_visa_hires']}")
        lines.append(f"  continuations: {features['continuations']}")
        lines.append(f"  amendments  : {features['amendments']}")
        lines.append(f"  total approvals: {features['total_approvals']}")
        if features.get("trend"):
            lines.append(f"  trend       : {features['trend']}")
    lines.append(f"  TIER {tier} -- {reason}")
    return "\n".join(lines)


def main(argv: list[str]) -> int:
    if len(argv) >= 2 and argv[1] in ("--load", "--reload"):
        n = ingest(force=True)
        print(f"Loaded {n} (company, fiscal_year) rows into {DB_PATH}")
        return 0

    n = ingest(force=False)
    if n:
        print(f"[ingest] populated DB with {n} (company, fiscal_year) rows")

    queries = argv[1:] if len(argv) > 1 else DEMO_NAMES
    for q in queries:
        feats = get_company_features(q)
        tier, reason = score_company(feats)
        print(_format_result(q, feats, tier, reason))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
